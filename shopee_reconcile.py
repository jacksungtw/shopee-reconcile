# -*- coding: utf-8 -*-
"""
Shopee 月結對帳自動化工具

功能：
1. 讀取一或多個 Shopee Order.completed Excel 檔（自動解密）
2. 篩選指定月份發票（含 typo 容錯：1153.X / 113.3.X 修正為 115.3.X）
3. 對照紙本帳逐日比對
4. 產出 3 份報表：
   - 月份對帳表 (含逐日對帳 / 摘要 / typo 清單)
   - 差異日訂單明細 (差>3000 的日子)
   - 工程師獎金統計 (沿用 v0.6.3 邏輯)

用法：
  方法 1：把 Excel 拖到 對帳工具.bat 上
  方法 2：python shopee_reconcile.py <檔1.xlsx> [檔2.xlsx ...] [--month=3] [--year=115]

紙本資料：
  在 Excel 同目錄放 紙本對帳_115年3月.csv（格式見 _紙本範例.csv），會自動讀入比對。
  沒有也沒關係，會跳過紙本對帳。
"""
import os
import sys
import re
import io
import argparse
from datetime import datetime
from pathlib import Path

# 強制 stdout/stderr 用 UTF-8（避免 Windows cp950 編碼錯誤）
if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    except Exception:
        pass

import pandas as pd

# 嘗試載入解密
try:
    import msoffcrypto
    HAS_DECRYPT = True
except ImportError:
    HAS_DECRYPT = False

# 從環境變數讀密碼；找不到就讀 .env；都沒有就用空字串（檔案沒加密時沒影響）
def _load_password():
    pw = os.environ.get("SHOPEE_PASSWORD", "").strip()
    if pw:
        return pw
    env_file = Path(__file__).parent / ".env"
    if env_file.exists():
        for line in env_file.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if line.startswith("SHOPEE_PASSWORD"):
                _, _, v = line.partition("=")
                return v.strip().strip('"').strip("'")
    return ""


DEFAULT_PASSWORD = _load_password()

# 必要欄位
COL_ORDER_ID = "訂單編號"
COL_STATUS = "訂單狀態"
COL_QTY = "數量"
COL_BUYER_PAID = "買家總支付金額"
COL_NOTE = "備註"
COL_RETURN_QTY = "退貨數量"

VALID_CODES = {"B", "E", "J", "K", "P", "S"}
SALES_BONUS_RATE = 0.01
REPAIR_BONUS_PER_UNIT = 80


def read_excel_auto(path: str, password: str = DEFAULT_PASSWORD) -> pd.DataFrame:
    """自動處理加密與否的 Excel 讀取"""
    if HAS_DECRYPT:
        try:
            with open(path, "rb") as fh:
                of = msoffcrypto.OfficeFile(fh)
                if of.is_encrypted():
                    of.load_key(password=password)
                    buf = io.BytesIO()
                    of.decrypt(buf)
                    buf.seek(0)
                    return pd.read_excel(buf, engine="openpyxl")
        except Exception as e:
            print(f"[WARN] 解密失敗，改用標準讀取：{e}")
    return pd.read_excel(path, engine="openpyxl")


def parse_invoice_day(note, target_year=115, target_month=3):
    """
    從備註中抽取目標月份的「日」（含 typo 容錯）
    回傳: int day 或 None
    支援格式：
      標準: 115.3.5  / 115. 3.5 / 115.3. 5
      typo: 1153.5  -> 解讀為 115.3.5
      typo: 113.3.5 (僅當有 ZM 開頭發票號) -> 解讀為 115.3.5
    """
    if pd.isna(note):
        return None
    s = str(note)
    # 標準格式 (允許空白)
    pat = rf"{target_year}\s*\.\s*{target_month}\s*\.\s*(\d+)"
    m = re.search(pat, s.replace(" ", ""))
    if m:
        return int(m.group(1))
    # typo 1: 1153.X -> 115.3.X (僅當 month=3)
    if target_month == 3:
        m = re.match(rf"^\s*{target_year}{target_month}\.(\d+)", s)
        if m:
            return int(m.group(1))
        # typo 2: 113.3.X with ZM
        m = re.match(rf"113\.{target_month}\.(\d+)", s)
        if m and "ZM" in s:
            return int(m.group(1))
    return None


def detect_typo(note, target_year=115, target_month=3):
    """偵測 typo 並回傳 (錯誤類型, 建議修正)，沒有則回傳 None"""
    if pd.isna(note):
        return None
    s = str(note)
    if target_month == 3:
        if re.match(rf"^\s*{target_year}{target_month}\.\d", s):
            fixed = re.sub(rf"^(\s*){target_year}{target_month}\.",
                          rf"\g<1>{target_year}.{target_month}.", s, count=1)
            return (f"{target_year}{target_month}.X 應為 {target_year}.{target_month}.X", fixed)
        if re.match(rf"113\.{target_month}\.\d", s) and "ZM" in s:
            fixed = re.sub(rf"^113\.{target_month}\.",
                          f"{target_year}.{target_month}.", s, count=1)
            return (f"113.{target_month}.X 應為 {target_year}.{target_month}.X", fixed)
    return None


def load_paper_data(input_dir: Path, year: int, month: int) -> dict:
    """讀取紙本對帳 CSV（若存在）。回傳 {day: amount}"""
    candidates = [
        input_dir / f"紙本對帳_{year}年{month}月.csv",
        input_dir / f"paper_{year}_{month:02d}.csv",
    ]
    for path in candidates:
        if path.exists():
            try:
                df = pd.read_csv(path, encoding="utf-8-sig")
                if "日" in df.columns and "金額" in df.columns:
                    return dict(zip(df["日"].astype(int), df["金額"].astype(int)))
            except Exception as e:
                print(f"[WARN] 紙本資料讀取失敗：{e}")
    return {}


def reconcile(input_files: list, target_year: int, target_month: int,
              output_dir: Path, paper_data: dict = None):
    """主對帳流程"""
    paper_data = paper_data or {}

    # ==== 1. 讀檔合併 ====
    print(f"\n[1/4] 讀取 {len(input_files)} 個 Excel 檔案...")
    dfs = []
    for f in input_files:
        print(f"  - {os.path.basename(f)}")
        d = read_excel_auto(f)
        d["_source_file"] = os.path.basename(f)
        dfs.append(d)
    df = pd.concat(dfs, ignore_index=True)
    print(f"  總共 {len(df)} 筆原始資料")

    # 必要欄位檢查
    for c in [COL_ORDER_ID, COL_STATUS, COL_BUYER_PAID, COL_NOTE]:
        if c not in df.columns:
            raise ValueError(f"找不到必要欄位：{c}")

    # 數值清理
    df[COL_BUYER_PAID] = pd.to_numeric(df[COL_BUYER_PAID], errors="coerce").fillna(0)
    if COL_RETURN_QTY in df.columns:
        df[COL_RETURN_QTY] = pd.to_numeric(df[COL_RETURN_QTY], errors="coerce").fillna(0)

    # ==== 2. 篩選目標月份 ====
    print(f"\n[2/4] 篩選 {target_year}年{target_month}月 發票...")
    done = df[df[COL_STATUS].astype(str).str.startswith("已完成")].copy()
    done = done.drop_duplicates(subset=[COL_ORDER_ID], keep="first")
    done["_invoice_day"] = done[COL_NOTE].apply(
        lambda v: parse_invoice_day(v, target_year, target_month))
    target = done[done["_invoice_day"].notna()].copy()
    target["_invoice_day"] = target["_invoice_day"].astype(int)
    print(f"  符合 {target_year}.{target_month}.X 訂單：{len(target)} 筆")

    # ==== 3. 產出對帳表 ====
    print(f"\n[3/4] 產出對帳表...")
    daily_agg = target.groupby("_invoice_day").agg(
        Excel筆數=(COL_ORDER_ID, "count"),
        Excel金額=(COL_BUYER_PAID, "sum")
    ).reset_index().rename(columns={"_invoice_day": "日"})

    all_days = sorted(set(daily_agg["日"].tolist()) | set(paper_data.keys()))
    rows = []
    for d in all_days:
        row = daily_agg[daily_agg["日"] == d]
        cnt = int(row["Excel筆數"].iloc[0]) if not row.empty else 0
        amt = float(row["Excel金額"].iloc[0]) if not row.empty else 0
        p = paper_data.get(d, 0)
        diff = amt - p if p else 0
        if not paper_data:
            status = ""
        elif abs(diff) > 3000:
            status = "★大"
        elif abs(diff) > 500:
            status = "★"
        else:
            status = "OK"
        rows.append({
            "日期": f"{target_month}月{d}日",
            "Excel筆數": cnt,
            "Excel金額": amt,
            "紙本金額": p,
            "差額": diff,
            "狀態": status,
        })
    daily_df = pd.DataFrame(rows)
    total_e = daily_df["Excel金額"].sum()
    total_p = daily_df["紙本金額"].sum()
    total_row = pd.DataFrame([{
        "日期": "合計",
        "Excel筆數": daily_df["Excel筆數"].sum(),
        "Excel金額": total_e,
        "紙本金額": total_p,
        "差額": total_e - total_p,
        "狀態": "",
    }])
    daily_full = pd.concat([daily_df, total_row], ignore_index=True)

    # typo 清單
    typo_rows = []
    for _, r in target.iterrows():
        t = detect_typo(r[COL_NOTE], target_year, target_month)
        if t:
            typo_rows.append({
                "訂單編號": r[COL_ORDER_ID],
                "金額": r[COL_BUYER_PAID],
                "錯誤備註": str(r[COL_NOTE]),
                "建議修正": t[1],
                "錯誤類型": t[0],
            })
    typo_df = pd.DataFrame(typo_rows) if typo_rows else pd.DataFrame(
        columns=["訂單編號", "金額", "錯誤備註", "建議修正", "錯誤類型"])

    # 摘要
    summary_rows = []
    if paper_data:
        summary_rows.extend([
            {"項目": "紙本 合計", "金額": total_p, "備註": "★ 對帳目標"},
            {"項目": "Excel 合計", "金額": int(total_e), "備註": ""},
            {"項目": "差額（Excel - 紙本）", "金額": int(total_e - total_p),
             "備註": f"吻合度 {(1-abs(total_e-total_p)/total_p)*100:.1f}%" if total_p else ""},
        ])
    else:
        summary_rows.append({"項目": "Excel 合計", "金額": int(total_e), "備註": "（無紙本可對）"})
    summary_rows.append({"項目": "", "金額": None, "備註": ""})
    summary_rows.append({"項目": "目標月份", "金額": None,
                        "備註": f"{target_year}年{target_month}月"})
    summary_rows.append({"項目": "已完成訂單筆數", "金額": len(target), "備註": ""})
    summary_rows.append({"項目": "其中含退貨", "金額": int((target[COL_RETURN_QTY] > 0).sum()) if COL_RETURN_QTY in target.columns else None,
                        "備註": ""})
    summary_rows.append({"項目": "Typo 修正筆數", "金額": len(typo_rows), "備註": ""})
    summary_df = pd.DataFrame(summary_rows)

    # 寫入主對帳檔
    out_main = output_dir / f"對帳表_{target_year}年{target_month}月.xlsx"
    out_main = safe_path(out_main)
    with pd.ExcelWriter(out_main, engine="openpyxl") as w:
        daily_full.to_excel(w, sheet_name="逐日對帳", index=False)
        summary_df.to_excel(w, sheet_name="摘要", index=False)
        typo_df.to_excel(w, sheet_name="備註typo修正清單", index=False)
    style_workbook(out_main)
    print(f"  [OK] {out_main}")

    # ==== 差異日明細 ====
    if paper_data:
        big_diff = [d for d in all_days
                    if d in paper_data
                    and abs((float(daily_agg[daily_agg["日"] == d]["Excel金額"].iloc[0])
                            if d in daily_agg["日"].values else 0) - paper_data[d]) > 3000]
        if big_diff:
            cols = [c for c in [COL_ORDER_ID, "訂單成立日期", "訂單完成時間",
                                "實際出貨時間", COL_NOTE, COL_BUYER_PAID,
                                COL_RETURN_QTY, COL_STATUS] if c in target.columns]
            out_diff = output_dir / f"差異日明細_{target_year}年{target_month}月.xlsx"
            out_diff = safe_path(out_diff)
            with pd.ExcelWriter(out_diff, engine="openpyxl") as w:
                for d in big_diff:
                    sub = target[target["_invoice_day"] == d][cols].copy()
                    sub.to_excel(w, sheet_name=f"{target_month}月{d}日", index=False)
            print(f"  [OK] {out_diff} ({len(big_diff)} 個差異日)")

    # ==== 4. 工程師獎金 ====
    print(f"\n[4/4] 計算工程師獎金...")
    bonus_df = compute_bonus(target)
    if not bonus_df.empty:
        out_bonus = output_dir / f"工程師獎金_{target_year}年{target_month}月.xlsx"
        out_bonus = safe_path(out_bonus)
        bonus_df.to_excel(out_bonus, index=False)
        style_workbook(out_bonus)
        print(f"  [OK] {out_bonus}")

    # ==== 完成摘要 ====
    print("\n" + "=" * 60)
    print(f"[完成] 對帳完成！{target_year}年{target_month}月")
    print("=" * 60)
    print(f"訂單筆數：{len(target)}")
    print(f"Excel 合計：{total_e:,.0f} 元")
    if paper_data:
        print(f"紙本合計：{total_p:,.0f} 元")
        ratio = (1-abs(total_e-total_p)/total_p)*100 if total_p else 0
        print(f"差額：{total_e-total_p:+,.0f} 元 (吻合度 {ratio:.2f}%)")
    if typo_rows:
        print(f"Typo 修正：{len(typo_rows)} 筆 -> 已列在「備註typo修正清單」分頁")
    print(f"\n[輸出目錄] {output_dir}")


# ==== 工程師獎金解析引擎（v0.6.3 精確版）====


def is_valid_note(s):
    """只要看起來有工程師分配語法，就視為可解析候選"""
    if s is None:
        return False
    t = str(s).strip()
    if not t:
        return False
    return ("=" in t) or bool(re.search(r"[A-Za-z]\w*\s*\d", t))


def clean_alloc_segment(seg):
    """清理分配段，移除尾端非工程師代碼文字"""
    seg = seg.strip()
    m = re.match(
        r'^([A-Za-z]\w*\s*=\s*\d+(?:\s*[*]\s*\d+)?'
        r'(?:\s*[,，]\s*[A-Za-z]\w*\s*=\s*\d+(?:\s*[*]\s*\d+)?)*)',
        seg)
    if m:
        return m.group(1).strip()
    return seg


def extract_alloc_segment_single(note):
    """單行備註的分配段提取"""
    if note is None:
        return ""
    s = str(note).strip()
    if not s:
        return ""
    parts_bar = [p.strip() for p in s.split("|")]
    if len(parts_bar) >= 3:
        return clean_alloc_segment(parts_bar[1])
    elif len(parts_bar) == 2:
        left, right = parts_bar[0], parts_bar[1]

        def _is_alloc(text):
            if "=" in text and re.search(r'[A-Za-z]\w*\s*=\s*\d', text):
                return True
            if re.search(r'[A-Za-z]\w*\s*\d', text):
                return True
            return False

        left_is = _is_alloc(left)
        right_is = _is_alloc(right)
        if left_is and not right_is:
            return clean_alloc_segment(left)
        elif right_is and not left_is:
            return clean_alloc_segment(right)
        elif left_is and right_is:
            return clean_alloc_segment(right)
        else:
            left_letter = bool(re.search(r'[A-Za-z]', left))
            right_letter = bool(re.search(r'[A-Za-z]', right))
            if left_letter and not right_letter:
                return clean_alloc_segment(left)
            elif right_letter and not left_letter:
                return clean_alloc_segment(right)
            else:
                return clean_alloc_segment(left)
    else:
        return clean_alloc_segment(s)


def extract_alloc_segment(note):
    """從備註中提取工程師分配段（支援多行備註）"""
    if note is None:
        return ""
    s = str(note).strip()
    if not s:
        return ""
    lines = s.split("\n")
    if len(lines) >= 2:
        segs = []
        for line in lines:
            line = line.strip()
            if not line:
                continue
            ls = extract_alloc_segment_single(line)
            if ls:
                segs.append(ls)
        return ",".join(segs) if segs else ""
    else:
        return extract_alloc_segment_single(s)


def parse_note(note):
    """
    解析備註中的工程師分配（v0.6.3 完整引擎）

    回傳: (mode, parts) 或 (None, None)
      mode: "A"=支數比例 / "B"=固定金額
      parts: [{code, qty, amount}]
    """
    if note is None:
        return None, None

    alloc = extract_alloc_segment(note)
    if not alloc:
        return None, None

    if ("=" not in alloc) and (not re.search(r"[A-Za-z]\w*\s*\d", alloc)):
        return None, None

    # Token: 逗號分割再空格分割
    raw_tokens = [t.strip() for t in re.split(r"[,，]", alloc) if t.strip()]
    tokens = []
    for t in raw_tokens:
        sub = re.split(r"\s+", t)
        if len(sub) >= 2 and all(
            re.search(r'[A-Za-z]\s*=', s) or re.match(r'[A-Za-z]\s*\d', s)
            for s in sub
        ):
            tokens.extend(sub)
        else:
            tokens.append(t)

    parts = []
    has_money = False

    for t in tokens:
        # 容錯1: S1*130 -> S=1*130
        if "=" not in t:
            m0 = re.match(
                r"^([A-Za-z]\w*)\s*([0-9]+(?:\.[0-9]+)?)\s*\*\s*([0-9]+(?:\.[0-9]+)?)$",
                t)
            if m0:
                t = f"{m0.group(1)}={m0.group(2)}*{m0.group(3)}"
            else:
                # 容錯2: S600 -> S=600
                if "*" not in t:
                    m00 = re.match(
                        r"^([A-Za-z]\w*?)([0-9]+(?:\.[0-9]+)?)$", t)
                    if m00:
                        code_part = m00.group(1)
                        num_part = m00.group(2)
                        if num_part and code_part and len(code_part) > 0:
                            t = f"{code_part}={num_part}"

        # 正規解析
        m = re.match(
            r"^([A-Za-z]\w*)\s*=\s*([0-9]+(?:\.[0-9]+)?)"
            r"(?:\s*\*\s*([0-9]+(?:\.[0-9]+)?))?$",
            t)
        if not m:
            continue

        code = m.group(1).upper()
        if code not in VALID_CODES:
            continue
        v1 = float(m.group(2))
        v2 = m.group(3)

        if v2 is not None:
            qty = int(round(v1))
            unit = float(v2)
            has_money = True
            parts.append({"code": code, "qty": qty, "amount": qty * unit})
        else:
            if v1 <= 20 and abs(v1 - round(v1)) < 1e-9:
                parts.append({"code": code, "qty": int(v1), "amount": None})
            else:
                has_money = True
                parts.append({"code": code, "qty": 1, "amount": float(v1)})

    if not parts:
        return None, None

    return ("B" if has_money else "A"), parts


def compute_bonus(target_df):
    """工程師獎金計算（v0.6.3 精確解析引擎）"""
    rows = []
    processed = set()

    for _, r in target_df.iterrows():
        oid = r[COL_ORDER_ID]
        if oid in processed:
            continue
        if COL_RETURN_QTY in target_df.columns and r.get(COL_RETURN_QTY, 0) > 0:
            processed.add(oid)
            continue

        mode, parts = parse_note(r[COL_NOTE])
        if mode is None:
            processed.add(oid)
            continue

        buyer = float(r[COL_BUYER_PAID])

        if mode == "A":
            total_qty = sum(p["qty"] for p in parts)
            if total_qty <= 0:
                processed.add(oid)
                continue
            for p in parts:
                amt = buyer * p["qty"] / total_qty
                rows.append({
                    "工程師": p["code"],
                    "支數": p["qty"],
                    "金額": round(amt, 2),
                })
        else:
            for p in parts:
                amt = float(p["amount"] or 0.0)
                rows.append({
                    "工程師": p["code"],
                    "支數": p["qty"],
                    "金額": round(amt, 2),
                })

        processed.add(oid)

    if not rows:
        return pd.DataFrame()

    df = pd.DataFrame(rows).groupby("工程師", as_index=False).agg(
        支數=("支數", "sum"), 金額=("金額", "sum"))
    df["銷售獎金_1%"] = (df["金額"] * SALES_BONUS_RATE).round(2)
    df["維修獎金_80元每支"] = (df["支數"] * REPAIR_BONUS_PER_UNIT).round(2)
    df["總獎金"] = (df["銷售獎金_1%"] + df["維修獎金_80元每支"]).round(2)
    return df


def safe_path(p: Path) -> Path:
    """避免覆蓋已開啟的檔案"""
    if not p.exists():
        return p
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return p.with_name(f"{p.stem}_{ts}{p.suffix}")


def style_workbook(path):
    """套用基本樣式"""
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = load_workbook(path)
        thin = Side(border_style="thin", color="888888")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = PatternFill("solid", start_color="1F4E78")
        header_font = Font(name="Microsoft JhengHei", bold=True, color="FFFFFF", size=11)
        body_font = Font(name="Microsoft JhengHei", size=11)
        big_fill = PatternFill("solid", start_color="F4B084")
        star_fill = PatternFill("solid", start_color="FFE699")
        total_fill = PatternFill("solid", start_color="D9E1F2")

        for sn in wb.sheetnames:
            ws = wb[sn]
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.font = body_font
                    cell.border = border
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = "#,##0;(#,##0);-"
                        cell.alignment = Alignment(horizontal="right")
                # 標記列
                status_cell = next((c for c in row if c.value in ("★大", "★", "OK")), None)
                if status_cell:
                    if status_cell.value == "★大":
                        for c in row: c.fill = big_fill
                    elif status_cell.value == "★":
                        for c in row: c.fill = star_fill
                if row[0].value == "合計":
                    for c in row:
                        c.fill = total_fill
                        c.font = Font(name="Microsoft JhengHei", bold=True, size=11)
            for i, col in enumerate(ws.columns, 1):
                max_len = max((len(str(c.value or "")) for c in col), default=10)
                ws.column_dimensions[get_column_letter(i)].width = min(max(max_len * 1.3 + 2, 12), 50)
            ws.row_dimensions[1].height = 28
        wb.save(path)
    except Exception as e:
        print(f"  [WARN] 樣式套用失敗（不影響資料）：{e}")


def main():
    parser = argparse.ArgumentParser(description="Shopee 月結對帳工具")
    parser.add_argument("files", nargs="+", help="Order.completed Excel 檔案路徑")
    parser.add_argument("--month", type=int, default=None, help="目標月份（民國年月）預設自動偵測")
    parser.add_argument("--year", type=int, default=115, help="民國年（預設 115）")
    parser.add_argument("--out", type=str, default=None, help="輸出目錄（預設與輸入同目錄）")
    args = parser.parse_args()

    files = [os.path.abspath(f) for f in args.files]
    for f in files:
        if not os.path.exists(f):
            print(f"[ERROR] 檔案不存在：{f}")
            sys.exit(1)

    # 自動推測月份：用檔名中的 YYYYMM 抓最大月份
    if args.month is None:
        months = []
        for f in files:
            m = re.search(r"(\d{4})(\d{2})\d{2}", os.path.basename(f))
            if m:
                months.append(int(m.group(2)))
        if months:
            args.month = max(months)
            print(f"[INFO] 自動偵測月份：{args.month}月")
        else:
            args.month = int(input("請輸入要對帳的月份（例如 3）："))

    out_dir = Path(args.out) if args.out else Path(files[0]).parent
    out_dir.mkdir(parents=True, exist_ok=True)

    paper = load_paper_data(out_dir, args.year, args.month)
    if paper:
        print(f"[INFO] 找到紙本對帳資料：{len(paper)} 天")
    else:
        print("[INFO] 無紙本對帳資料（在輸入目錄放 紙本對帳_115年X月.csv 可自動比對）")

    reconcile(files, args.year, args.month, out_dir, paper)


def _pause():
    try:
        input("\n按 Enter 結束...")
    except (EOFError, OSError):
        pass


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"\n[ERROR] {e}")
        import traceback
        traceback.print_exc()
        _pause()
        sys.exit(1)
    _pause()
